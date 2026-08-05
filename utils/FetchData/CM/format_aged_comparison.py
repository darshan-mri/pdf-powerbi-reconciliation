"""
Script to format aged comparison results into a more readable Excel format.
Creates a formatted output with proper styling, grouping by Tenant,
and highlighting for mismatched values.
"""
import sys
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(_PROJECT_ROOT) not in sys.path:
    sys.path.append(str(_PROJECT_ROOT))

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from utils.config_util import Config
from datetime import datetime


def format_aged_comparison(input_file: str = None, output_file: str = None,
                           period: str = None, format_all_sheets: bool = True):
    """
    Format the aged comparison results into a more readable Excel format.

    Args:
        input_file: Path to the input comparison results Excel file
        output_file: Path for the formatted output Excel file
        period: Period string (e.g., "12/25"). If None, uses current month/year
        format_all_sheets: If True, formats all sheets. If False, only main comparison
    """
    # Load config
    config = Config()

    # Set default file paths from config if not provided
    if input_file is None:
        input_file = config.get('CM.AGED', 'ComparisonResult')
    if output_file is None:
        output_file = str(input_file).replace('.xlsx', '_formatted.xlsx')

    # Set default period to current month/year if not provided
    if period is None:
        now = datetime.now()
        period = now.strftime("%m/%y")

    print(f"[INFO] Formatting aged comparison: {input_file}")

    # Read the comparison results - check which sheets exist
    xl_file = pd.ExcelFile(input_file)
    available_sheets = xl_file.sheet_names

    print(f"[INFO] Available sheets: {available_sheets}")

    # Read the main comparison sheet
    df = pd.read_excel(input_file)

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create formatted sheet
    print(f"[INFO] Formatting comparison data...")
    ws = wb.create_sheet(title="AGED - Comparison")

    # Format the sheet
    _format_sheet(ws, df, period)

    # Optionally create a summary sheet
    if format_all_sheets:
        print(f"[INFO] Creating summary sheet...")
        create_summary_sheet(wb, df)

    # Save the workbook
    wb.save(output_file)
    print(f"[SUCCESS] Formatted aged comparison saved to: {output_file}")
    return output_file


def _format_sheet(ws, df, period):
    """
    Format a single worksheet with aged comparison data.

    Args:
        ws: openpyxl worksheet
        df: DataFrame with comparison data
        period: Period string
    """
    # Define styles
    header_font = Font(bold=True, size=11)
    subheader_font = Font(bold=True, size=10)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')
    left_align = Alignment(horizontal='left')

    # Add header rows
    ws['A1'] = f"Aged Report Comparison - Period {period}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    # Add column headers starting at row 3
    headers = ['Invoice ID', 'Tenant', 'Aged Report Element', 'Aged Report Value',
               'PMX Report Element', 'PMX Report Value', 'Matched']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = light_gray_fill

    # Process data - transform to the new format
    current_row = 4

    # Define the aging buckets to compare
    aging_fields = [
        ("Total", "Total"),
        ("Current", "Current"),
        ("Month_1", "30 Days"),
        ("Month_2", "60 Days"),
        ("Month_3", "90 Days"),
        ("Month_4", "120+ Days")
    ]

    # Group by Tenant (each tenant gets multiple rows for different aging buckets)
    for _, row in df.iterrows():
        invoice_id_extracted = row.get('InvoiceID_extracted', '')
        invoice_id_report = row.get('InvoiceID_report', '')
        invoice_id = invoice_id_extracted if pd.notna(invoice_id_extracted) else invoice_id_report

        tenant_extracted = row.get('Tenant_extracted', '')
        tenant_report = row.get('Tenant_report', '')
        tenant = tenant_extracted if pd.notna(tenant_extracted) else tenant_report

        # Track if this is the first row for this invoice (for Invoice ID display)
        is_first_row = True

        # Write each aging field comparison
        for field_name, display_name in aging_fields:
            # Get values
            extracted_value = row.get(f'{field_name}_extracted', 0)
            report_value = row.get(f'{field_name}_report', 0)
            is_match = row.get(f'{field_name}_match', False)

            # Format values
            extracted_formatted = format_value(extracted_value)
            report_formatted = format_value(report_value)

            # Write Invoice ID only on first row
            if is_first_row:
                cell_invoice = ws.cell(row=current_row, column=1, value=invoice_id)
                cell_invoice.alignment = left_align
                is_first_row = False

            # Write Tenant (only on first aging field row)
            if field_name == "Total":
                cell_tenant = ws.cell(row=current_row, column=2, value=tenant)
                cell_tenant.alignment = left_align

            # Column C - Aged Report Element
            cell_c = ws.cell(row=current_row, column=3, value=display_name)
            cell_c.alignment = right_align

            # Column D - Aged Report Value
            cell_d = ws.cell(row=current_row, column=4, value=report_formatted)
            cell_d.alignment = right_align

            # Column E - PMX Report Element
            cell_e = ws.cell(row=current_row, column=5, value=display_name)
            cell_e.alignment = center_align

            # Column F - PMX Report Value
            cell_f = ws.cell(row=current_row, column=6, value=extracted_formatted)
            cell_f.alignment = right_align

            # Column G - Matched
            match_text = "Match" if is_match else "No Match"
            cell_g = ws.cell(row=current_row, column=7, value=match_text)
            cell_g.alignment = center_align

            # Apply yellow highlighting if not matched
            if not is_match:
                cell_g.fill = yellow_fill
                # Also highlight the mismatched values
                cell_d.fill = yellow_fill
                cell_f.fill = yellow_fill

            current_row += 1

        # Add blank row between invoices for readability
        current_row += 1

    # Adjust column widths
    column_widths = [15, 35, 22, 18, 22, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Freeze header rows
    ws.freeze_panes = 'A4'


def format_value(value):
    """Format numeric values with comma separators and 2 decimal places."""
    if pd.isna(value):
        return ""
    try:
        num_value = float(value)
        return f"{num_value:,.2f}"
    except (ValueError, TypeError):
        return str(value)


def create_summary_sheet(wb, df):
    """Create a summary sheet with totals and statistics."""
    ws_summary = wb.create_sheet("Summary", 0)  # Insert as first sheet

    # Define styles
    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)

    # Calculate statistics
    total_records = len(df)

    # Count matches for each field
    aging_fields = ["Total", "Current", "Month_1", "Month_2", "Month_3", "Month_4"]

    # Write summary header
    ws_summary['A1'] = "Aged Report Comparison Summary"
    ws_summary['A1'].font = header_font
    ws_summary.merge_cells('A1:B1')

    # Basic stats
    ws_summary['A3'] = "Total Records"
    ws_summary['A3'].font = subheader_font
    ws_summary['B3'] = total_records

    # Overall match rate
    if 'Overall_Match' in df.columns:
        overall_matched = df['Overall_Match'].sum()
        match_rate = (overall_matched / total_records * 100) if total_records > 0 else 0

        ws_summary['A4'] = "Overall Matched"
        ws_summary['B4'] = overall_matched

        ws_summary['A5'] = "Overall Match Rate"
        ws_summary['B5'] = f"{match_rate:.1f}%"

    # Field-level match statistics
    ws_summary['A7'] = "Field-Level Match Statistics"
    ws_summary['A7'].font = subheader_font
    ws_summary.merge_cells('A7:C7')

    ws_summary['A8'] = "Field"
    ws_summary['B8'] = "Matched"
    ws_summary['C8'] = "Match Rate"

    for i, col in enumerate(ws_summary['A8':'C8'][0]):
        col.font = Font(bold=True)

    current_row = 9
    field_display_names = {
        "Total": "Total",
        "Current": "Current",
        "Month_1": "30 Days",
        "Month_2": "60 Days",
        "Month_3": "90 Days",
        "Month_4": "120+ Days"
    }

    for field_name, display_name in field_display_names.items():
        match_col = f"{field_name}_match"
        if match_col in df.columns:
            matched_count = df[match_col].sum()
            match_rate = (matched_count / total_records * 100) if total_records > 0 else 0

            ws_summary.cell(row=current_row, column=1, value=display_name)
            ws_summary.cell(row=current_row, column=2, value=matched_count)
            ws_summary.cell(row=current_row, column=3, value=f"{match_rate:.1f}%")

            current_row += 1

    # Variance statistics
    ws_summary.cell(row=current_row + 1, column=1, value="Total Variance Analysis").font = subheader_font
    ws_summary.merge_cells(f'A{current_row + 1}:C{current_row + 1}')

    current_row += 2

    ws_summary.cell(row=current_row, column=1, value="Field")
    ws_summary.cell(row=current_row, column=2, value="Total Extracted")
    ws_summary.cell(row=current_row, column=3, value="Total Report")
    ws_summary.cell(row=current_row, column=4, value="Variance")

    for col in ws_summary[f'A{current_row}':f'D{current_row}'][0]:
        col.font = Font(bold=True)

    current_row += 1

    for field_name, display_name in field_display_names.items():
        extracted_col = f"{field_name}_extracted"
        report_col = f"{field_name}_report"

        if extracted_col in df.columns and report_col in df.columns:
            total_extracted = df[extracted_col].sum()
            total_report = df[report_col].sum()
            variance = total_extracted - total_report

            ws_summary.cell(row=current_row, column=1, value=display_name)
            ws_summary.cell(row=current_row, column=2, value=f"{total_extracted:,.2f}")
            ws_summary.cell(row=current_row, column=3, value=f"{total_report:,.2f}")
            ws_summary.cell(row=current_row, column=4, value=f"{variance:,.2f}")

            current_row += 1

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 18
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 18


if __name__ == "__main__":
    # Run the formatter with default settings
    try:
        print("="*80)
        print("AGED COMPARISON FORMATTER")
        print("="*80)
        print()

        output_path = format_aged_comparison()

        print()
        print("="*80)
        print("[SUCCESS] FORMATTING COMPLETE")
        print("="*80)
        print(f"[OUTPUT] Output file: {output_path}")
        print()
        print("[TIP] Open the file in Excel to review the formatted comparison!")

    except Exception as e:
        print(f"[ERROR] {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


