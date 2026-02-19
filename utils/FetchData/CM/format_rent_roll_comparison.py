"""
Script to format rent roll comparison results into a more readable Excel format.
Creates a formatted output with proper styling, grouping by Building ID and Suite ID,
and highlighting for mismatched values.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from utils.config_util import Config
from datetime import datetime


def format_rent_roll_comparison(input_file: str = None, output_file: str = None,
                                  period: str = None, suite_sq_ft: str = "All"):
    """
    Format the rent roll comparison results into a more readable Excel format.

    Only includes these columns:
    - Building ID
    - Suite ID
    - GLA Sqft
    - Monthly Base Rent
    - Annual Rate PSF
    - Monthly Other Income

    Args:
        input_file: Path to the input comparison results Excel file
        output_file: Path for the formatted output Excel file
        period: Period string (e.g., "12/25"). If None, uses current month/year
        suite_sq_ft: Suite Sq Ft filter value (default: "All")
    """
    # Load config
    config = Config()

    # Set default file paths from config if not provided
    if input_file is None:
        input_file = config.get('CM.ROLL', 'ComparisonResult')
    if output_file is None:
        output_file = str(input_file).replace('.xlsx', '_formatted.xlsx')

    # Set default period to current month/year if not provided
    if period is None:
        now = datetime.now()
        period = now.strftime("%m/%y")

    print(f"📊 Reading rent roll comparison from: {input_file}")

    # Read the comparison results
    df = pd.read_excel(input_file)

    # Remove any "APPLIED FILTERS" rows
    if len(df) > 0 and 'Building ID' in df.columns:
        filter_mask = df['Building ID'].astype(str).str.contains('APPLIED FILTERS', case=False, na=False)
        rows_before = len(df)
        df = df[~filter_mask].copy()
        df = df.reset_index(drop=True)
        rows_removed = rows_before - len(df)
        if rows_removed > 0:
            print(f"🧹 Removed {rows_removed} APPLIED FILTERS row(s)")

    print(f"📄 Loaded {len(df)} records")
    print(f"📄 Available columns: {list(df.columns)}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Create worksheet
    ws = wb.create_sheet(title="RENT ROLL")

    # Count unique combinations before formatting
    if 'Building ID' in df.columns and 'Suite ID' in df.columns:
        original_count = len(df.groupby(['Building ID', 'Suite ID']))
        print(f"📊 Unique Building ID + Suite ID combinations: {original_count}")
    else:
        original_count = len(df)
        print(f"📊 Total records: {original_count}")

    # Format the sheet
    formatted_count = _format_sheet(ws, df, period, suite_sq_ft)

    # Verify counts match
    print(f"📊 Formatted groups: {formatted_count}")
    if formatted_count == original_count:
        print(f"✅ Count verification passed: {original_count} records formatted")
    else:
        print(f"⚠️ Count mismatch: Original={original_count}, Formatted={formatted_count}")

    # Save the workbook
    wb.save(output_file)
    print(f"✅ Formatted rent roll comparison saved to: {output_file}")
    return output_file


def _format_sheet(ws, df, period, suite_sq_ft):
    """
    Format a single worksheet with comparison data.

    Args:
        ws: openpyxl worksheet
        df: DataFrame with comparison data
        period: Period string
        suite_sq_ft: Suite Sq ft value

    Returns:
        int: Number of unique Building ID + Suite ID groups formatted
    """
    # Define styles
    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')

    # Add header rows
    ws['A1'] = f"Period {period}"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Suite Sq ft - {suite_sq_ft}"
    ws['A2'].font = Font(bold=True)

    # Add column headers starting at row 3
    headers = ['Building ID', 'Suite ID', 'Insights Element', 'Insights Value',
               'PMX Report Element', 'PMX Report Value', 'Matched']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align

    # Define the columns to compare
    columns_to_compare = [
        ('GLA Sqft', 'GLA Sqft'),
        ('Monthly Base Rent', 'Monthly Base Rent'),
        ('Annual Rate PSF', 'Annual Rate PSF'),
        ('Monthly Other Income', 'Monthly Other Income')
    ]

    # Process data - transform to the new format
    current_row = 4
    formatted_groups = 0

    # Group by Building ID and Suite ID
    if 'Building ID' in df.columns and 'Suite ID' in df.columns:
        grouped = df.groupby(['Building ID', 'Suite ID'], dropna=False)

        for (building_id, suite_id), group_df in grouped:
            # Get the first row from this group (should only be one)
            row = group_df.iloc[0]

            # Write each comparison field
            first_row_in_group = True

            for insights_col, pmx_col in columns_to_compare:
                # Column names with suffixes
                extracted_col = f"{insights_col}_Extracted"
                report_col = f"{insights_col}_Report"
                match_col = f"{insights_col}_Match"

                # Get values
                insights_value = format_value(row.get(report_col))
                pmx_value = format_value(row.get(extracted_col))
                matched = row.get(match_col, True)  # Default to True if column doesn't exist

                # Write Building ID and Suite ID only on first row for this record
                if first_row_in_group:
                    building_cell = ws.cell(row=current_row, column=1, value=str(building_id) if pd.notna(building_id) else "")
                    building_cell.alignment = right_align
                    suite_cell = ws.cell(row=current_row, column=2, value=str(suite_id) if pd.notna(suite_id) else "")
                    suite_cell.alignment = right_align
                    first_row_in_group = False
                else:
                    ws.cell(row=current_row, column=1, value="")
                    ws.cell(row=current_row, column=2, value="")

                # Write data row
                write_data_row(ws, current_row,
                              insights_element=insights_col,
                              insights_value=insights_value,
                              pmx_element=pmx_col,
                              pmx_value=pmx_value,
                              matched=matched,
                              yellow_fill=yellow_fill,
                              right_align=right_align)

                current_row += 1

            formatted_groups += 1

    else:
        print("⚠️ Warning: Required columns 'Building ID' or 'Suite ID' not found in data")

    # Adjust column widths
    column_widths = [15, 12, 20, 15, 20, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    return formatted_groups


def format_value(value):
    """Format numeric values with comma separators and 2 decimal places."""
    if pd.isna(value):
        return ""
    try:
        num_value = float(value)
        return f"{num_value:,.2f}"
    except (ValueError, TypeError):
        return str(value)


def write_data_row(ws, row, insights_element, insights_value, pmx_element,
                   pmx_value, matched, yellow_fill, right_align):
    """Write a single data row with appropriate styling."""

    # Column C - Insights Element
    cell_c = ws.cell(row=row, column=3, value=insights_element)
    cell_c.alignment = Alignment(horizontal='right')

    # Column D - Insights Value
    cell_d = ws.cell(row=row, column=4, value=insights_value)
    cell_d.alignment = right_align

    # Column E - PMX Report Element
    cell_e = ws.cell(row=row, column=5, value=pmx_element)
    cell_e.alignment = Alignment(horizontal='center')

    # Column F - PMX Report Value
    cell_f = ws.cell(row=row, column=6, value=pmx_value)
    cell_f.alignment = right_align

    # Column G - Matched
    match_text = "Match" if matched else "No Match"
    cell_g = ws.cell(row=row, column=7, value=match_text)
    cell_g.alignment = right_align

    # Apply yellow highlighting if not matched
    if not matched:
        cell_g.fill = yellow_fill


def create_summary_sheet(wb, df):
    """Create a summary sheet with totals and statistics."""
    ws_summary = wb.create_sheet("Summary")

    # Calculate statistics
    total_records = len(df)

    # Count matches for each field
    fields = ['GLA Sqft', 'Monthly Base Rent', 'Annual Rate PSF', 'Monthly Other Income']
    match_cols = [f"{field}_Match" for field in fields]

    # Write summary
    ws_summary['A1'] = "Rent Roll Comparison Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    summary_data = [("Total Records", total_records)]

    for field in fields:
        match_col = f"{field}_Match"
        if match_col in df.columns:
            matched_count = df[match_col].sum()
            match_rate = (matched_count / total_records * 100) if total_records > 0 else 0
            summary_data.append((f"{field} Matched", matched_count))
            summary_data.append((f"{field} Match Rate", f"{match_rate:.1f}%"))

    # Overall match
    if 'Overall_Match' in df.columns:
        overall_matched = df['Overall_Match'].sum()
        overall_rate = (overall_matched / total_records * 100) if total_records > 0 else 0
        summary_data.append(("Overall Matched", overall_matched))
        summary_data.append(("Overall Match Rate", f"{overall_rate:.1f}%"))

    for i, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 15


if __name__ == "__main__":
    # Run the formatter with default settings
    format_rent_roll_comparison()


