"""
Script to format ledger comparison results into a more readable Excel format.
Creates a formatted output with proper styling, grouping by Building ID,
and highlighting for mismatched values.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from utils.config_util import Config
from datetime import datetime


def format_ledger_comparison(input_file: str = None, output_file: str = None,
                              period: str = None, suite_sq_ft: str = "All",
                              format_all_sheets: bool = True):
    """
    Format the ledger comparison results into a more readable Excel format.

    Args:
        input_file: Path to the input comparison results Excel file
        output_file: Path for the formatted output Excel file
        period: Period string (e.g., "09/25"). If None, uses current month/year
        suite_sq_ft: Suite Sq Ft filter value (default: "All")
        format_all_sheets: If True, formats all comparison sheets. If False, only LeaseComparison
    """
    # Load config
    config = Config()

    # Set default file paths from config if not provided
    if input_file is None:
        input_file = config.get('CM.LEDGER', 'ComparisonResult')
    if output_file is None:
        output_file = str(input_file).replace('.xlsx', '_formatted.xlsx')

    # Set default period to current month/year if not provided
    if period is None:
        now = datetime.now()
        period = now.strftime("%m/%y")

    # Read the comparison results - check which sheets exist
    xl_file = pd.ExcelFile(input_file)
    available_sheets = xl_file.sheet_names

    print(f"📊 Available sheets: {available_sheets}")

    # Create a new workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    sheets_to_format = []

    if format_all_sheets:
        # Format all comparison sheets
        if "LeaseComparison" in available_sheets:
            sheets_to_format.append(("LeaseComparison", "LEDGER - Lease"))
        if "BuildingComparison" in available_sheets:
            sheets_to_format.append(("BuildingComparison", "LEDGER - Building"))
    else:
        # Format only LeaseComparison
        if "LeaseComparison" in available_sheets:
            sheets_to_format.append(("LeaseComparison", "LEDGER"))

    if not sheets_to_format:
        raise ValueError(f"No comparison sheets found in {input_file}")

    # Format each sheet
    for sheet_name, tab_name in sheets_to_format:
        print(f"📄 Formatting sheet: {sheet_name}")
        df = pd.read_excel(input_file, sheet_name=sheet_name)

        # Determine if this is a building-level comparison (no 'lease' column)
        is_building_level = 'lease' not in df.columns

        # Create worksheet
        ws = wb.create_sheet(title=tab_name)

        # Format the sheet
        _format_sheet(ws, df, period, suite_sq_ft, is_building_level)

    # Save the workbook
    wb.save(output_file)
    print(f"✅ Formatted ledger comparison saved to: {output_file}")
    return output_file


def _format_sheet(ws, df, period, suite_sq_ft, is_building_level=False):
    """
    Format a single worksheet with comparison data.

    Args:
        ws: openpyxl worksheet
        df: DataFrame with comparison data
        period: Period string
        suite_sq_ft: Suite Sq ft value
        is_building_level: True if this is building-level data (no lease column)
    """
    # Define styles
    header_font = Font(bold=True)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')

    # Add header rows
    ws['A1'] = f"Period {period}"
    ws['A1'].font = Font(bold=True)
    ws['A2'] = f"Suite Sq ft - {suite_sq_ft}"
    ws['A2'].font = Font(bold=True)

    # Add column headers starting at row 3
    # Use "Lease Id" for lease-level data, "Building Id" for building-level data
    first_column_header = 'Building Id' if is_building_level else 'Lease Id'
    headers = [first_column_header, 'Insights Element', 'Insights Value',
               'PMX Report Element', 'PMX Report Value', 'Matched']

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_align

    # Process data - transform to the new format
    current_row = 4

    # Group by building
    for building in df['building'].unique():
        building_df = df[df['building'] == building]

        # For each record in this building
        for idx, (_, row) in enumerate(building_df.iterrows()):
            # Determine the identifier (lease for lease-level, building for building-level)
            if is_building_level:
                identifier = building
            else:
                lease = row.get('lease')
                identifier = lease if pd.notna(lease) else building

            # Format values with 2 decimal places if numeric
            charges_extracted = format_value(row.get('charges_extracted'))
            cash_extracted = format_value(row.get('cash_receipts_extracted'))
            charges_report = format_value(row.get('charges_report'))
            cash_report = format_value(row.get('cash_receipts_report'))

            charges_match = row.get('charges_match', False)
            cash_match = row.get('cash_match', False)

            # Write Building/Lease ID only on first row for this record
            building_cell = ws.cell(row=current_row, column=1, value=identifier)
            building_cell.alignment = right_align

            # Write Total Billings row
            write_data_row(ws, current_row,
                          insights_element="Charges",
                          insights_value=charges_report,
                          pmx_element="Total Billings",
                          pmx_value=charges_extracted,
                          matched=charges_match,
                          yellow_fill=yellow_fill,
                          right_align=right_align)

            current_row += 1

            # Write Total Credits row
            ws.cell(row=current_row, column=1, value="")
            write_data_row(ws, current_row,
                          insights_element="Cash Receipts",
                          insights_value=cash_report,
                          pmx_element="Total Credits",
                          pmx_value=cash_extracted,
                          matched=cash_match,
                          yellow_fill=yellow_fill,
                          right_align=right_align)

            current_row += 1

    # Adjust column widths
    column_widths = [12, 18, 15, 20, 18, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width


def format_value(value):
    """Format numeric values with comma separators and 2 decimal places."""
    if pd.isna(value):
        return ""
    try:
        num_value = float(value)
        return f"{num_value:,.2f}"
    except (ValueError, TypeError):
        return str(value)


def write_data_row(ws, row, insights_element, insights_value, pmx_element,
                   pmx_value, matched, yellow_fill, right_align):
    """Write a single data row with appropriate styling."""

    # Column B - Insights Element
    cell_b = ws.cell(row=row, column=2, value=insights_element)
    cell_b.alignment = Alignment(horizontal='right')

    # Column C - Insights Value
    cell_c = ws.cell(row=row, column=3, value=insights_value)
    cell_c.alignment = right_align

    # Column D - PMX Report Element
    cell_d = ws.cell(row=row, column=4, value=pmx_element)
    cell_d.alignment = Alignment(horizontal='center')

    # Column E - PMX Report Value
    cell_e = ws.cell(row=row, column=5, value=pmx_value)
    cell_e.alignment = right_align

    # Column F - Matched
    match_text = "Match" if matched else "No Match"
    cell_f = ws.cell(row=row, column=6, value=match_text)
    cell_f.alignment = right_align

    # Apply yellow highlighting if not matched
    if not matched:
        cell_f.fill = yellow_fill


def create_summary_sheet(wb, df):
    """Create a summary sheet with totals and statistics."""
    ws_summary = wb.create_sheet("Summary")

    # Calculate statistics
    total_records = len(df)
    charges_matched = df['charges_match'].sum() if 'charges_match' in df.columns else 0
    cash_matched = df['cash_match'].sum() if 'cash_match' in df.columns else 0
    overall_matched = df['Overall_Match'].sum() if 'Overall_Match' in df.columns else 0

    # Write summary
    ws_summary['A1'] = "Ledger Comparison Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)

    summary_data = [
        ("Total Records", total_records),
        ("Charges Matched", charges_matched),
        ("Charges Match Rate", f"{(charges_matched/total_records*100):.1f}%" if total_records > 0 else "N/A"),
        ("Cash Receipts Matched", cash_matched),
        ("Cash Match Rate", f"{(cash_matched/total_records*100):.1f}%" if total_records > 0 else "N/A"),
        ("Overall Matched", overall_matched),
        ("Overall Match Rate", f"{(overall_matched/total_records*100):.1f}%" if total_records > 0 else "N/A"),
    ]

    for i, (label, value) in enumerate(summary_data, 3):
        ws_summary.cell(row=i, column=1, value=label)
        ws_summary.cell(row=i, column=2, value=value)


if __name__ == "__main__":
    # Run the formatter with default settings
    format_ledger_comparison()
