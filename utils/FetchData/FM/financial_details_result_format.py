import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from utils.config_util import Config
import re

# ---------- CONFIG ----------
cfg = Config()
input_file = str(cfg.get("FM.VARIANCE", "ComparisonResult"))
output_file = str(cfg.get("FM.VARIANCE", "VarianceFormatted"))

# ---------- READ MTD + YTD ----------
mtd = pd.read_excel(input_file, sheet_name="MTD_Comparison")
ytd = pd.read_excel(input_file, sheet_name="YTD_Comparison")

# ---------- SECTION BUILDER ----------
def build_section(df, period_label):
    rows = []
    for _, row in df.iterrows():
        label = row["Label"]

        # MTD / YTD header row
        rows.append({
            "PMX Report Element": period_label,
            "PMX Value": "",
            "Dashboard Value": "",
            "Dashboard Element": "",
            "Match": ""
        })

        # CATEGORY header row
        rows.append({
            "PMX Report Element": label,
            "PMX Value": "",
            "Dashboard Value": "",
            "Dashboard Element": label,
            "Match": ""
        })

        # METRICS
        for metric in ["Actual", "Budget", "Variance", "Variance %"]:
            pmx_col = f"{metric} {period_label}_Extracted"
            dash_col = f"{metric} {period_label}_Report"

            pmx_val = row.get(pmx_col, "")
            dash_val = row.get(dash_col, "")

            if isinstance(pmx_val, (int, float)):
                pmx_val = f"{pmx_val:,.2f}"
            if isinstance(dash_val, (int, float)):
                dash_val = f"{dash_val:,.2f}"

            rows.append({
                "PMX Report Element": metric,
                "PMX Value": pmx_val,
                "Dashboard Value": dash_val,
                "Dashboard Element": "",
                "Match": "Match" if pmx_val == dash_val else "No Match"
            })
    return rows

# ---------- BUILD FULL FINANCIAL DETAILS ----------
output_rows = []
for lbl in mtd["Label"].unique():
    output_rows.extend(build_section(mtd[mtd["Label"] == lbl], "MTD"))
    output_rows.extend(build_section(ytd[ytd["Label"] == lbl], "YTD"))

final_df = pd.DataFrame(output_rows)

# ---------- SAVE ----------
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    final_df.to_excel(writer, index=False, sheet_name="Financial Details")

# ---------- LOAD FOR FORMATTING ----------
wb = load_workbook(output_file)
ws = wb["Financial Details"]

bold = Font(bold=True)
center = Alignment(horizontal="center")
gray = PatternFill("solid", fgColor="D9D9D9")
yellow = PatternFill("solid", fgColor="FFFF00")

# ---------- PERIOD HEADER ----------
def insert_period_header(ws, filename):
    m = re.search(r"_(\d{4})\.xlsx$", filename)
    if not m:
        return
    code = m.group(1)
    period = f"{code[:2]}/{code[2:]}"  # "0724" -> "07/24"

    ws.insert_rows(1, 2)
    ws["A1"] = f"Period: {period}"
    ws["A1"].font = bold

# Insert before any formatting offsets
insert_period_header(ws, input_file)

# ---------- FORMAT FINANCIAL DETAILS ----------
for row in ws.iter_rows(min_row=3, max_col=5):  # Start from row 3 (after period + blank)
    first = row[0]
    match_cell = row[4]
    text = str(first.value).upper()

    if text in ("MTD", "YTD"):
        for c in row:
            c.font = bold
            c.fill = gray
            c.alignment = center

    elif text and text.isupper() is False and not text.startswith("[[") and not text in ("MTD", "YTD"):
        # Category names already bold in build
        pass

    if match_cell.value == "No Match":
        match_cell.fill = yellow

# Autofit
for col in ws.columns:
    maxlen = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            maxlen = max(maxlen, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = maxlen + 2

# ===============================================================
# SUMMARY SHEET EXTRACTION
# ===============================================================

df_fd = pd.read_excel(output_file, sheet_name="Financial Details")

def extract_actual(keyword):
    results = []

    for i in range(len(df_fd) - 1):
        row_lbl = str(df_fd.loc[i, "PMX Report Element"])
        next_lbl = str(df_fd.loc[i+1, "PMX Report Element"])

        if keyword.lower() in row_lbl.lower() and next_lbl.lower() == "actual":

            period = None
            j = i
            while j >= 0:
                p = str(df_fd.loc[j, "PMX Report Element"]).upper()
                if p in ("MTD", "YTD"):
                    period = p
                    break
                j -= 1

            results.append({
                "period": period,
                "pmx": df_fd.loc[i+1, "PMX Value"],
                "dash": df_fd.loc[i+1, "Dashboard Value"]
            })

    return results

def create_summary_sheet(name, keywords):
    if name in wb.sheetnames:
        del wb[name]

    wsx = wb.create_sheet(name)

    # --- First insert the period header ---
    insert_period_header(wsx, input_file)

    # Header row (now row 3)
    wsx.append(["PMX Report Element", "PMX Value", "Dashboard Value", "Dashboard Element", "Match"])
    for c in wsx[3]:
        c.font = bold

    # --- MTD block ---
    wsx.append(["MTD", "", "", "", ""])
    for c in wsx[wsx.max_row]:
        c.font = bold
        c.fill = gray
        c.alignment = center

    for kw in keywords:
        items = extract_actual(kw)
        for item in items:
            if item["period"] == "MTD":
                wsx.append([kw, item["pmx"], item["dash"], kw, ""])
                row = wsx.max_row

                wsx[f"E{row}"] = f'=IF(B{row}=C{row},"Match","No Match")'
                if item["pmx"] != item["dash"]:
                    wsx[f"E{row}"].fill = yellow

    # --- YTD block ---
    wsx.append(["YTD", "", "", "", ""])
    for c in wsx[wsx.max_row]:
        c.font = bold
        c.fill = gray
        c.alignment = center

    for kw in keywords:
        items = extract_actual(kw)
        for item in items:
            if item["period"] == "YTD":
                wsx.append([kw, item["pmx"], item["dash"], kw, ""])
                row = wsx.max_row

                wsx[f"E{row}"] = f'=IF(B{row}=C{row},"Match","No Match")'
                if item["pmx"] != item["dash"]:
                    wsx[f"E{row}"].fill = yellow

    # Conditional formatting
    dxf = DifferentialStyle(fill=PatternFill("solid", fgColor="FFFF00"))
    rule = Rule(type="containsText", operator="containsText", text="No Match", dxf=dxf)
    wsx.conditional_formatting.add("E3:E1048576", rule)

    # Autofit
    for col in wsx.columns:
        maxlen = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                maxlen = max(maxlen, len(str(cell.value)))
        wsx.column_dimensions[col_letter].width = maxlen + 2

# -------- CREATE THE 2 SUMMARY SHEETS --------
create_summary_sheet("Financial NOI Analysis", [
    "Total Income",
    "TOTAL EXPENSES",
    "Net Operating Income"
])

create_summary_sheet("Financial Portfolio Hub", [
    "Total Income",
    "TOTAL EXPENSES",
    "Net Operating Income",
    "TOTAL CAPITAL EXPENDITURE",
    "Non Operating Expenses"
])

wb.save(output_file)
print("✅ Completed:", output_file)

