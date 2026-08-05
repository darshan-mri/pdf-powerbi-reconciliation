import re
import sys
from typing import Any, cast
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# Allow running this file directly (python utils/BPG/DetRollComparison.py).
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.append(str(PROJECT_ROOT))

from utils.config_util import Config

# ===========================================
# CONFIG
# ===========================================

cfg = Config()

exported_excel = cfg.get(section="BPG", key="DetRollExtracted")
extracted_excel = cfg.get(section="BPG", key="DetRollExcel")
output_excel = Path(extracted_excel).with_name("det_roll_resident_level_comparison.xlsx")

TOLERANCE = 0.01


def is_missing(value: object) -> bool:
    return bool(pd.isna(cast(Any, value)))


def is_present(value: object) -> bool:
    return bool(pd.notna(cast(Any, value)))


def normalize_text(value: object) -> str:
    """Normalize resident names (also handles 'Last,First' style)."""
    if is_missing(value):
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9,\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()

    if "," in text:
        last, first = text.split(",", 1)
        text = f"{first.strip()} {last.strip()}"

    tokens = re.findall(r"[a-z0-9]+", text)
    return "".join(tokens)


def to_first_last_display_name(value: object) -> str:
    """Convert names like 'Last,First' to 'First Last' for display and matching."""
    if is_missing(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""

    if "," in text:
        last, first = text.split(",", 1)
        text = f"{first.strip()} {last.strip()}"

    # Collapse multiple spaces and title-case for readability in outputs.
    text = re.sub(r"\s+", " ", text).strip()
    return text.title()


def join_unique(values: pd.Series) -> str:
    """Join unique values into a comma-separated string."""
    units = sorted(
        {str(v).strip() for v in values if is_present(v) and str(v).strip()}
    )
    return ", ".join(units)


def normalize_unit(value: object) -> str:
    """Normalize unit ids so formatting differences do not break joins."""
    if is_missing(value):
        return ""
    text = str(value).strip().upper()
    text = re.sub(r"\s+", "", text)
    return text


def to_money(value: object) -> float:
    if is_missing(value):
        return np.nan
    if isinstance(value, (int, float, np.number)):
        return float(value)
    text = str(value).strip()
    if not text:
        return np.nan
    neg = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").replace(",", "")
    try:
        num = float(text)
    except ValueError:
        return np.nan
    return -num if neg else num


def first_non_zero(*values: object) -> float:
    for value in values:
        if is_present(value):
            num = float(cast(Any, value))
            if abs(num) > TOLERANCE:
                return num
    return np.nan


def rewrite_det_roll_names_inplace(
    path: Path,
    sheet_name: str = "Rent Roll",
    name_column: str = "Name",
) -> tuple[int, Path]:
    """Rewrite det_roll resident names in-place from Last,First to First Last."""
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{sheet_name}' not found in {path}")

    ws = wb[sheet_name]
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    header_map = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(header_cells) if cell.value is not None}
    if name_column not in header_map:
        raise KeyError(f"Column '{name_column}' not found in sheet '{sheet_name}'")

    name_col_idx = header_map[name_column]
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=name_col_idx)
        original = "" if cell.value is None else str(cell.value)
        updated = to_first_last_display_name(original)
        if updated and updated != original.strip():
            cell.value = updated
            changed += 1

    try:
        wb.save(path)
        return changed, path
    except PermissionError:
        fallback = path.with_name(
            f"{path.stem}_names_updated_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
        )
        wb.save(fallback)
        return changed, fallback


def load_exported(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Export")
    cols = ["Property Code", "Building", "Unit ID", "Resident Name", "Expiring Rent"]
    df = df[cols].copy()
    df["Expiring Rent"] = df["Expiring Rent"].apply(to_money)
    df["resident_key"] = df["Resident Name"].apply(normalize_text)
    df["unit_key"] = df["Unit ID"].apply(normalize_unit)
    df = df[df["resident_key"] != ""]

    grouped = (
        df.groupby("resident_key", dropna=False, as_index=False)
        .agg(
            {
                "Property Code": "first",
                "Building": "first",
                "Unit ID": join_unique,
                "Resident Name": "first",
                "resident_key": "first",
                "unit_key": "first",
                "Expiring Rent": "sum",
            }
        )
        .rename(columns={"Resident Name": "Resident Name_Extracted", "Unit ID": "Units_Extracted"})
    )
    return grouped


def load_det_roll(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="Rent Roll")
    cols = [
        "Property",
        "Building",
        "Unit",
        "Name",
        "Market Rent",
        "Lease Rent",
        "Total Billing",
    ]
    df = df[cols].copy()
    df = df[df["Name"].notna() & (df["Name"].astype(str).str.strip() != "")]

    df = df.rename(columns={"Unit": "Unit ID", "Name": "Resident Name_DetRoll"})
    df["Resident Name_DetRoll"] = df["Resident Name_DetRoll"].apply(to_first_last_display_name)
    for col in ["Market Rent", "Lease Rent", "Total Billing"]:
        df[col] = df[col].apply(to_money)

    df["resident_key"] = df["Resident Name_DetRoll"].apply(normalize_text)
    df["unit_key"] = df["Unit ID"].apply(normalize_unit)
    df = df[df["resident_key"] != ""]

    grouped = (
        df.groupby("resident_key", dropna=False, as_index=False)
        .agg(
            {
                "Property": "first",
                "Building": "first",
                "Unit ID": join_unique,
                "Resident Name_DetRoll": "first",
                "resident_key": "first",
                "unit_key": "first",
                "Market Rent": "sum",
                "Lease Rent": "sum",
                "Total Billing": "sum",
            }
        )
        .rename(columns={"Unit ID": "Units_DetRoll"})
    )
    return grouped


def evaluate_strategy(df_both: pd.DataFrame, strategy_col: str) -> dict:
    valid = df_both[df_both[strategy_col].notna()].copy()
    if valid.empty:
        return {
            "Strategy": strategy_col,
            "Compared Rows": 0,
            "Exact/Within 0.01": 0,
            "Within 1.00": 0,
            "Mean Abs Diff": np.nan,
            "Median Abs Diff": np.nan,
        }

    abs_diff = (valid["Expiring Rent"] - valid[strategy_col]).abs()
    return {
        "Strategy": strategy_col,
        "Compared Rows": int(len(valid)),
        "Exact/Within 0.01": int((abs_diff <= TOLERANCE).sum()),
        "Within 1.00": int((abs_diff <= 1.00).sum()),
        "Mean Abs Diff": float(abs_diff.mean()),
        "Median Abs Diff": float(abs_diff.median()),
    }


def main() -> None:
    det_roll_path = Path(extracted_excel)

    if "--rewrite-det-roll-names" in sys.argv:
        print("Rewriting det_roll resident names in-place (Last,First -> First Last)...")
        changed, saved_path = rewrite_det_roll_names_inplace(det_roll_path)
        det_roll_path = saved_path
        print(f"Updated rows in 'Rent Roll' Name column: {changed}")
        if saved_path != Path(extracted_excel):
            print(f"det_roll.xlsx is open; saved updated copy to: {saved_path}")
        else:
            print(f"Saved updated file: {saved_path}")
        print()

    print("=" * 90)
    print("DET ROLL RESIDENT-LEVEL EXPIRING RENT COMPARISON")
    print("=" * 90)
    print(f"DetRollExtracted file : {exported_excel}")
    print(f"det_roll.xlsx file    : {det_roll_path} (sheet: Rent Roll)")

    left = load_exported(Path(exported_excel))
    right = load_det_roll(det_roll_path)

    comparison = left.merge(
        right,
        on="resident_key",
        how="outer",
        suffixes=("_Extracted", "_DetRoll"),
        indicator=True,
    )

    # Resident-level rule requested by user:
    # If Total Billing is zero/missing -> compare against Market Rent, else Total Billing.
    comparison["Rule_MarketIfBillingZero_ElseTotalBilling"] = np.where(
        comparison["Total Billing"].fillna(0).abs() <= TOLERANCE,
        comparison["Market Rent"],
        comparison["Total Billing"],
    )

    # Additional strategies to check what best aligns with Expiring Rent.
    comparison["Strategy_MarketRent"] = comparison["Market Rent"]
    comparison["Strategy_LeaseRent"] = comparison["Lease Rent"]
    comparison["Strategy_TotalBilling"] = comparison["Total Billing"]
    comparison["Strategy_Total_Else_Lease_Else_Market"] = comparison.apply(
        lambda row: first_non_zero(row["Total Billing"], row["Lease Rent"], row["Market Rent"]),
        axis=1,
    )
    comparison["Strategy_Lease_Else_Total_Else_Market"] = comparison.apply(
        lambda row: first_non_zero(row["Lease Rent"], row["Total Billing"], row["Market Rent"]),
        axis=1,
    )

    comparison["Selected Rent (Requested Rule)"] = comparison[
        "Rule_MarketIfBillingZero_ElseTotalBilling"
    ]
    comparison["Diff (Expiring - Selected)"] = (
        comparison["Expiring Rent"] - comparison["Selected Rent (Requested Rule)"]
    )

    comparison["Status"] = ""
    comparison.loc[comparison["_merge"] == "left_only", "Status"] = "Missing in det_roll Rent Roll"
    comparison.loc[comparison["_merge"] == "right_only", "Status"] = "Missing in DetRollExtracted"

    both_mask = comparison["_merge"] == "both"
    match_mask = both_mask & (
        comparison["Diff (Expiring - Selected)"].abs() <= TOLERANCE
    )
    comparison.loc[both_mask, "Status"] = "Mismatch"
    comparison.loc[match_mask, "Status"] = "Match"

    # Resident missing lists (resident name only, regardless of unit).
    extracted_resident_set = {
        str(v) for v in left["resident_key"].dropna().astype(str) if str(v).strip()
    }
    det_resident_set = {
        str(v) for v in right["resident_key"].dropna().astype(str) if str(v).strip()
    }

    missing_in_det = sorted(extracted_resident_set - det_resident_set)
    missing_in_extracted = sorted(det_resident_set - extracted_resident_set)

    resident_lookup_extracted = (
        left[["resident_key", "Resident Name_Extracted"]]
        .drop_duplicates(subset=["resident_key"])
        .set_index("resident_key")["Resident Name_Extracted"]
        .to_dict()
    )
    resident_lookup_det = (
        right[["resident_key", "Resident Name_DetRoll"]]
        .drop_duplicates(subset=["resident_key"])
        .set_index("resident_key")["Resident Name_DetRoll"]
        .to_dict()
    )

    missing_in_det_df = pd.DataFrame(
        {
            "resident_key": missing_in_det,
            "Resident Name (DetRollExtracted)": [
                resident_lookup_extracted.get(str(k), "") for k in missing_in_det
            ],
        }
    )
    missing_in_extracted_df = pd.DataFrame(
        {
            "resident_key": missing_in_extracted,
            "Resident Name (det_roll Rent Roll)": [
                resident_lookup_det.get(str(k), "") for k in missing_in_extracted
            ],
        }
    )

    both_rows = comparison[comparison["_merge"] == "both"].copy()
    strategy_cols = [
        "Rule_MarketIfBillingZero_ElseTotalBilling",
        "Strategy_MarketRent",
        "Strategy_LeaseRent",
        "Strategy_TotalBilling",
        "Strategy_Total_Else_Lease_Else_Market",
        "Strategy_Lease_Else_Total_Else_Market",
    ]
    strategy_summary = pd.DataFrame(
        [evaluate_strategy(both_rows, col) for col in strategy_cols]
    ).sort_values(
        by=["Exact/Within 0.01", "Within 1.00", "Mean Abs Diff"],
        ascending=[False, False, True],
    )

    # Keep resident-friendly columns first.
    ordered_cols = [
        "Status",
        "resident_key",
        "Resident Name_Extracted",
        "Resident Name_DetRoll",
        "Units_Extracted",
        "Units_DetRoll",
        "Expiring Rent",
        "Market Rent",
        "Lease Rent",
        "Total Billing",
        "Selected Rent (Requested Rule)",
        "Diff (Expiring - Selected)",
        "Rule_MarketIfBillingZero_ElseTotalBilling",
        "Strategy_MarketRent",
        "Strategy_LeaseRent",
        "Strategy_TotalBilling",
        "Strategy_Total_Else_Lease_Else_Market",
        "Strategy_Lease_Else_Total_Else_Market",
        "_merge",
    ]
    comparison = comparison[[c for c in ordered_cols if c in comparison.columns]]

    run_output_excel = det_roll_path.with_name("det_roll_resident_level_comparison.xlsx")
    run_output_excel.parent.mkdir(parents=True, exist_ok=True)
    written_output = run_output_excel
    try:
        with pd.ExcelWriter(run_output_excel, engine="openpyxl") as writer:
            comparison.to_excel(writer, sheet_name="Resident Comparison", index=False)
            missing_in_det_df.to_excel(writer, sheet_name="Missing in det_roll", index=False)
            missing_in_extracted_df.to_excel(writer, sheet_name="Missing in Extracted", index=False)
            strategy_summary.to_excel(writer, sheet_name="Strategy Match Summary", index=False)
    except PermissionError:
        # Fallback when the target workbook is open in Excel.
        fallback = run_output_excel.with_name(
            f"{run_output_excel.stem}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}{run_output_excel.suffix}"
        )
        with pd.ExcelWriter(fallback, engine="openpyxl") as writer:
            comparison.to_excel(writer, sheet_name="Resident Comparison", index=False)
            missing_in_det_df.to_excel(writer, sheet_name="Missing in det_roll", index=False)
            missing_in_extracted_df.to_excel(writer, sheet_name="Missing in Extracted", index=False)
            strategy_summary.to_excel(writer, sheet_name="Strategy Match Summary", index=False)
        written_output = fallback
        print(f"\nPrimary output file is open. Saved to fallback file: {fallback}")

    print(f"\nSaved comparison to: {written_output}")


if __name__ == "__main__":
    main()
